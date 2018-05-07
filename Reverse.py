import xlrd
import geocoder
import openpyxl
from openpyxl import load_workbook
location = "C:\\Users\\ashish.baboo\\Desktop\\Cleaning Tools\\Reversed Template.xlsx"
workbook = xlrd.open_workbook(location)
sheet = workbook.sheet_by_index(0)


Rows = sheet.nrows
Cols = sheet.ncols
data = [[sheet.cell_value(r,c) for c in range(Cols)] for r in range(Rows)]
for i in range(1,Rows):
    print (i)
    Lat = data[i][9]
    Long = data[i][10]
    print(Lat,Long)
    g = geocoder.google([Lat,Long],method='reverse',language='en' , key = "AIzaSyD4yJPQqmCBR9m18z9iMVME-2_T7Nt0Ln8")
    print (g.housenumber,",",g.route,",",g.city,",",g.county,",",g.state,",",g.postal,",",g.country)
    wb = load_workbook('Reversed Template.xlsx')
    ws1 = wb.get_sheet_by_name("Locations")
    r = ws1.cell(row = i+1, column = 1)
    r.value = i
    c = ws1.cell(row = i+1, column = 3)
    c.value = g.housenumber
    d = ws1.cell(row = i+1, column = 4)
    d.value = g.route
    e = ws1.cell(row = i+1, column = 5)
    e.value = g.city
    f = ws1.cell(row = i+1, column = 6)
    f.value = g.county
    h = ws1.cell(row = i+1, column = 7)
    h.value = g.state_long
    k = ws1.cell(row = i+1, column = 8)
    k.value = g.postal
    h = ws1.cell(row = i+1, column = 9)
    h.value = g.country_long
    y = ws1.cell(row = i+1, column = 2)
    y.value = g.address
    wb.save('Reversed Template.xlsx')
